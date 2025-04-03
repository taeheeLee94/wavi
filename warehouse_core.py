import openpyxl
from collections import deque
import os
import itertools
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

def load_grid(filename):
    wb = openpyxl.load_workbook(filename, data_only=True)
    ws = wb.active
    grid = {}
    start_cells = []
    max_row = ws.max_row
    max_col = ws.max_column

    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            coord = cell.coordinate
            val = cell.value
            if val is not None and isinstance(val, str):
                val_str = val.strip()
                if val_str == "START":
                    cell_type = "START"
                elif val_str and val_str[0].isupper():
                    cell_type = "Corridor"
                elif val_str and val_str[0].islower():
                    cell_type = "warehouse"
                else:
                    cell_type = "Corridor"
            else:
                cell_type = "Corridor"
            grid[(cell.row, cell.column)] = {
                'coord': coord,
                'type': cell_type,
                'value': val
            }
            if cell_type == "START":
                start_cells.append((cell.row, cell.column))
    return grid, start_cells, max_row, max_col

def coordinate_to_index(coord_str):
    letter, number = coordinate_from_string(coord_str)
    col = column_index_from_string(letter)
    return (number, col)

def get_neighbors(cell, grid):
    r, c = cell
    neighbors = []
    for dr, dc in [(1,0), (-1,0), (0,1), (0,-1)]:
        nxt = (r+dr, c+dc)
        if nxt in grid and grid[nxt]['type'] in ["Corridor", "START"]:
            neighbors.append(nxt)
    return neighbors

def bfs(start, goal, grid):
    queue = deque([start])
    came_from = {start: None}
    while queue:
        current = queue.popleft()
        if current == goal:
            path = []
            while current is not None:
                path.append(current)
                current = came_from[current]
            path.reverse()
            return path
        for nxt in get_neighbors(current, grid):
            if nxt not in came_from:
                came_from[nxt] = current
                queue.append(nxt)
    return None

def find_warehouse_cell(warehouse_str, grid):
    for (r, c), info in grid.items():
        if info['value'] is not None and str(info['value']).strip() == warehouse_str and info['type'] == "warehouse":
            return (r, c)
    return None

def find_adjacent_corridor(warehouse_cell, grid):
    if not warehouse_cell:
        return None
    r, c = warehouse_cell
    for dc in [-1, 1]:  # 좌우 우선
        adj = (r, c+dc)
        if adj in grid and grid[adj]['type'] in ["Corridor", "START"]:
            return adj
    return None

def compute_all_pairwise(nodes, grid):
    pairwise = {}
    n = len(nodes)
    for i in range(n):
        for j in range(i+1, n):
            path = bfs(nodes[i], nodes[j], grid)
            dist = len(path) if path else float('inf')
            pairwise[(i, j)] = (dist, path)
            pairwise[(j, i)] = (dist, list(reversed(path)) if path else None)
    return pairwise

def solve_tsp(nodes, pairwise):
    n = len(nodes)
    if n == 1:
        return [0, 0], 0
    best_order, best_cost = None, float('inf')
    for perm in itertools.permutations(range(1, n)):
        order = [0] + list(perm) + [0]
        cost, valid = 0, True
        for i in range(len(order)-1):
            d, _ = pairwise.get((order[i], order[i+1]), (float('inf'), None))
            if d == float('inf'):
                valid = False
                break
            cost += d
        if valid and cost < best_cost:
            best_order, best_cost = order, cost
    return best_order, best_cost

def reconstruct_full_route(order, nodes, pairwise):
    full_route = []
    for i in range(len(order)-1):
        _, path = pairwise[(order[i], order[i+1])]
        if path is None:
            return None
        if i > 0:
            path = path[1:]
        full_route.extend(path)
    return full_route